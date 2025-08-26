import oracledb
import pandas as pd
import os
import logging
import concurrent.futures
<<<<<<< HEAD
from dotenv import load_dotenv 
from datetime import datetime
from tqdm import tqdm
import csv
=======
from openpyxl.styles import PatternFill
>>>>>>> bb3374c8ebbc2839fd2ae95848523d2e15acb503

# --- Configuração Colorama ---
try:
    from colorama import Fore, Style, init
    init(autoreset=True)
    GREEN = Fore.GREEN
    YELLOW = Fore.YELLOW
    RED = Fore.RED
    CYAN = Fore.CYAN
    MAGENTA = Fore.MAGENTA
    RESET = Style.RESET_ALL
except ImportError:
    print("Colorama não encontrado. Saída colorida desativada.")
    GREEN = YELLOW = RED = CYAN = MAGENTA = RESET = ""

# --- Constantes ---
LOG_FILE = "hardwarePDV.log" 
TABLE_NAME = "CONSINCO.BAR_HARDWARE_PDV"
CSV_FILE = "hardwarePDV_output.csv"
HARDWARE_FIELDS = ["PLACA_MAE", "PROCESSADOR", "CORES_THREADS", "RAM", "DISCO", "ARMAZENAMENTO", "RELEASE"]
PDV_INFO_FIELDS = ["IP", "SEGMENTO", "OPERACAO"]

# --- Configuração de Logging Principal ---
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO, 
    format="%(asctime)s - %(levelname)s - [%(name)s:%(funcName)s:%(lineno)d] - %(message)s"
)
logger = logging.getLogger("SCRIPT_PRINCIPAL") 

<<<<<<< HEAD
# --- Funções de Configuração ---
def load_env_configs():
    """Carrega e valida as configurações do .env para o script principal."""
    load_dotenv() 
    required_vars = {
        "ORACLE_USER": "seu_usuario_oracle",
        "ORACLE_PASSWORD": "sua_senha_oracle",
        "ORACLE_HOST": "host_oracle",
        "ORACLE_PORT": "porta_oracle",
        "ORACLE_SERVICE": "service_name_oracle",
        "SSH_USERNAME": "seu_usuario_ssh",
        "SSH_PASSWORD": "sua_senha_ssh"
=======
# Adicionar logger específico para validação de dados
validation_logger = logging.getLogger('validation')
validation_logger.setLevel(logging.INFO)
# Garantir que os logs de validação também vão para o arquivo principal
validation_logger.addHandler(logging.FileHandler("hardware_scan.log"))

# Configurações globais
CONNECTION_TIMEOUT = 8  # Timeout reduzido para acelerar falhas
COMMAND_TIMEOUT = 10    # Timeout para execução de comandos
MAX_WORKERS = 20        # Número máximo de threads paralelas
MAX_RETRIES = 2         # Número de tentativas para cada IP

class SSHClient:
    def __init__(self, ip, username, password, timeout=CONNECTION_TIMEOUT):
        self.ip = ip
        self.username = username
        self.password = password
        self.timeout = timeout
        self.client = None

    def connect(self):
        """Estabelece a conexão SSH"""
        if self.client is not None:
            return True
        self.client = paramiko.SSHClient()
        self.client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            self.client.connect(
                self.ip, 
                username=self.username, 
                password=self.password, 
                timeout=self.timeout,
                allow_agent=False,
                look_for_keys=False
            )
            return True
        except Exception as e:
            logging.error(f"Erro conectando a {self.ip}: {str(e)}")
            self.close()
            return False

    def execute_command(self, command, timeout=COMMAND_TIMEOUT):
        """Executa um comando SSH com timeout"""
        if self.client is None and not self.connect():
            return None
        try:
            # Verificar se self.client ainda é None após tentativa de conexão
            if self.client is None:
                return None
            stdin, stdout, stderr = self.client.exec_command(command, timeout=timeout)
            output = stdout.read().decode("utf-8", errors="ignore").strip()
            error = stderr.read().decode("utf-8", errors="ignore").strip()
            if error and "command not found" in error:
                return None
            return output
        except Exception as e:
            logging.error(f"Erro executando comando em {self.ip}: {str(e)}")
            return None

    def close(self):
        """Fecha a conexão SSH"""
        if self.client:
            try:
                self.client.close()
                self.client = None
            except Exception as e:
                logging.warning(f"Erro fechando conexão com {self.ip}: {str(e)}")

def obter_iso_release(cliente):
    """
    Obtém a versão do kernel usando o comando 'uname -r'.
    Retorna apenas a parte principal da versão (ex.: 3.15, 4.15, 5.3).
    Ignora sufixos como "-generic".
    """
    comando = "uname -r"
    resultado = cliente.execute_command(comando)
    if resultado:
        resultado = resultado.strip()
        try:
            # Remover sufixos como "-generic"
            versao_base = resultado.split('-')[0]
            # Extrair até o segundo número (ex.: "3.15")
            partes = versao_base.split('.')
            if len(partes) >= 2:
                return f"{partes[0]}.{partes[1]}"
            return versao_base  # Caso não tenha formato esperado
        except Exception as e:
            logging.error(f"Erro ao processar versão do kernel '{resultado}': {str(e)}")
            return "Versão desconhecida"
    return "Não detectado"

def obter_fabricante_placa_mae(cliente):
    """
    Obtém o fabricante da placa-mãe.
    """
    comandos = [
        "dmidecode -t baseboard | grep 'Manufacturer' | cut -d ':' -f2 | tr -s ' '",
        "cat /sys/devices/virtual/dmi/id/board_vendor",
        "lshw -c motherboard | grep 'vendor:' | cut -d ':' -f2 | tr -s ' '",
        "inxi -M | grep 'Mobo:' | cut -d ':' -f2 | cut -d ' ' -f2"
    ]
    for comando in comandos:
        resultado = cliente.execute_command(comando)
        if resultado:
            return resultado.strip()
    return "Não detectado"

def obter_modelo_placa_mae(cliente):
    """
    Obtém o modelo da placa-mãe.
    """
    comandos = [
        "dmidecode -t baseboard | grep 'Product' | cut -d ':' -f2 | tr -s ' '",
        "cat /sys/devices/virtual/dmi/id/board_name",
        "lshw -c motherboard | grep 'product:' | cut -d ':' -f2 | tr -s ' '",
        "inxi -M | grep 'Model:' | cut -d ':' -f2 | tr -s ' '"
    ]
    for comando in comandos:
        resultado = cliente.execute_command(comando)
        if resultado:
            return resultado.strip()
    return "Não detectado"

def obter_info_processador(cliente):
    """
    Obtém informações detalhadas do processador com detecção melhorada de núcleos/threads.
    """
    # Modelo do processador
    comandos_modelo = [
        "cat /proc/cpuinfo | grep 'model name' | head -1 | cut -d ':' -f2 | tr -s ' '",
        "lscpu | grep 'Model name' | cut -d ':' -f2 | tr -s ' '",
        "inxi -C | grep 'model name' | cut -d ':' -f2 | tr -s ' '"
    ]
    modelo = "Não detectado"
    for comando in comandos_modelo:
        resultado = cliente.execute_command(comando)
        if resultado:
            modelo = resultado.strip()
            break

    # Método melhorado para obter núcleos físicos e threads
    comandos_nucleos_fisicos = [
        "lscpu | grep 'Core(s) per socket' | awk '{print $4}'",
        "grep 'cpu cores' /proc/cpuinfo | uniq | awk '{print $4}'",
        "cat /sys/devices/system/cpu/cpu*/topology/core_id | sort -u | wc -l"
    ]
    comandos_sockets = [
        "lscpu | grep 'Socket(s)' | awk '{print $2}'",
        "dmidecode -t processor | grep 'Socket Designation' | wc -l"
    ]
    comandos_threads_totais = [
        "nproc",
        "grep -c ^processor /proc/cpuinfo",
        "lscpu | grep '^CPU(s):' | awk '{print $2}'"
    ]

    # Obter núcleos físicos por socket
    nucleos_por_socket = "1"  # valor padrão
    for comando in comandos_nucleos_fisicos:
        resultado = cliente.execute_command(comando)
        if resultado and resultado.strip().isdigit():
            nucleos_por_socket = resultado.strip()
            break

    # Obter número de sockets
    sockets = "1"  # valor padrão
    for comando in comandos_sockets:
        resultado = cliente.execute_command(comando)
        if resultado and resultado.strip().isdigit():
            sockets = resultado.strip()
            break

    # Calcular núcleos físicos totais
    nucleos_fisicos = (
        str(int(nucleos_por_socket) * int(sockets))
        if nucleos_por_socket.isdigit() and sockets.isdigit()
        else "1"
    )

    # Obter threads totais
    threads_total = "Não detectado"
    for comando in comandos_threads_totais:
        resultado = cliente.execute_command(comando)
        if resultado and resultado.strip().isdigit():
            threads_total = resultado.strip()
            break

    # Cálculo final de núcleos/threads
    nucleos_threads = (
        f"{nucleos_fisicos}/{threads_total}" 
        if nucleos_fisicos.isdigit() and threads_total.isdigit() 
        else "Não detectado"
    )
    return {
        "modelo": modelo,
        "nucleos_threads": nucleos_threads
>>>>>>> bb3374c8ebbc2839fd2ae95848523d2e15acb503
    }
    
    missing_vars = [key for key in required_vars if not os.getenv(key)]
    if missing_vars:
        error_msg = f"Variáveis ausentes no .env: {', '.join(missing_vars)}"
        logger.critical(error_msg)
        print(f"{RED}Erro: {error_msg}")
        print(f"{YELLOW}Adicione ao .env na raiz do projeto:")
        for var in missing_vars:
            print(f"{CYAN}  {var}=<{required_vars[var]}>")
        exit(1)

    configs = {key: os.getenv(key) for key in required_vars}
    configs["DSN"] = f"{configs['ORACLE_HOST']}:{configs['ORACLE_PORT']}/{configs['ORACLE_SERVICE']}"

    try:
        configs["MAX_WORKERS"] = int(os.getenv("MAX_WORKERS", "20"))
        configs["SAVE_INTERVAL"] = int(os.getenv("SAVE_INTERVAL", "10"))
    except ValueError as e:
        logger.warning(f"Erro ao ler MAX_WORKERS/SAVE_INTERVAL do .env: {e}. Usando padrões.")
        configs["MAX_WORKERS"] = 20
        configs["SAVE_INTERVAL"] = 10
        print(f"{YELLOW}Aviso: Erro ao ler MAX_WORKERS/SAVE_INTERVAL. Usando padrões.")

    logger.info(f"Configurações para script principal: MAX_WORKERS={configs['MAX_WORKERS']}, SAVE_INTERVAL={configs['SAVE_INTERVAL']}")
    return configs

# --- Funções de Banco ---
def check_table_exists(cursor):
    """Verifica se a tabela de hardware existe no banco."""
    try:
        cursor.execute(f"SELECT 1 FROM {TABLE_NAME} WHERE 1=0")
        logger.info(f"Tabela {TABLE_NAME} encontrada.")
        return "TABLE_EXISTS"
    except oracledb.DatabaseError as e:
        error_obj, = e.args
        if error_obj.code == 942: 
            logger.info(f"Tabela {TABLE_NAME} não encontrada.")
            return "TABLE_NOT_FOUND"
        logger.error(f"Erro DB ao verificar tabela {TABLE_NAME}: {str(e)} (Código: {error_obj.code})", exc_info=True)
        return "DB_ERROR"
    except Exception as e: 
        logger.error(f"Erro inesperado ao verificar tabela {TABLE_NAME}: {str(e)}", exc_info=True)
        return "DB_ERROR"

def create_oracle_table(cursor):
    """Cria a tabela de hardware no banco de dados Oracle."""
    create_table_sql = f"""
    CREATE TABLE {TABLE_NAME} (
        IP VARCHAR2(45), NROEMPRESA NUMBER(5) NOT NULL, NROCHECKOUT NUMBER(5) NOT NULL,
        SEGMENTO VARCHAR2(100), OPERACAO VARCHAR2(50), PLACA_MAE VARCHAR2(255),
        PROCESSADOR VARCHAR2(255), CORES_THREADS VARCHAR2(50), RAM VARCHAR2(50),
        DISCO VARCHAR2(50), ARMAZENAMENTO VARCHAR2(50), RELEASE VARCHAR2(100),
        STATUS VARCHAR2(10), DTAINCLUSAO DATE, DTAATUALIZACAO DATE,
        CONSTRAINT PK_BAR_HARDWARE_PDV PRIMARY KEY (NROEMPRESA, NROCHECKOUT)
    )"""
    try:
        cursor.execute(create_table_sql)
        logger.info(f"Tabela {TABLE_NAME} criada com sucesso.")
        print(f"{GREEN}Tabela {TABLE_NAME} criada com sucesso.")
        return True
    except Exception as e: 
        logger.error(f"Erro ao criar tabela {TABLE_NAME}: {str(e)}", exc_info=True)
        print(f"{RED}Erro ao criar tabela {TABLE_NAME}: {str(e)}")
        return False

def get_existing_data(cursor, nroempresa, nrocheckout):
    """Busca dados de hardware existentes para um PDV específico."""
    try:
        cursor.execute(
            f"SELECT IP, NROEMPRESA, NROCHECKOUT, SEGMENTO, OPERACAO, PLACA_MAE, PROCESSADOR, CORES_THREADS, RAM, DISCO, ARMAZENAMENTO, RELEASE, STATUS, DTAINCLUSAO, DTAATUALIZACAO FROM {TABLE_NAME} WHERE NROEMPRESA = :1 AND NROCHECKOUT = :2",
            (nroempresa, nrocheckout)
        )
        result = cursor.fetchone()
        if result:
            columns = [desc[0].upper() for desc in cursor.description]
            return dict(zip(columns, result))
        return None
    except Exception as e:
        logger.error(f"Erro ao buscar dados existentes para PDV {nroempresa}-{nrocheckout}: {str(e)}", exc_info=True)
        return None

def has_relevant_data_changes(new_data_dict, existing_data_db_dict):
    """Verifica alterações nos campos de dados (IP, Info PDV, Hardware)."""
    if not existing_data_db_dict: 
        return True 
    
    fields_to_compare = PDV_INFO_FIELDS + HARDWARE_FIELDS
    for field in fields_to_compare:
        new_value = str(new_data_dict.get(field, "Não detectado")).strip()
        existing_value = str(existing_data_db_dict.get(field, "Não detectado")).strip()
        if new_value != existing_value:
            logger.info(f"Mudança de DADOS no PDV {new_data_dict.get('NROEMPRESA')}-{new_data_dict.get('NROCHECKOUT')}: Campo '{field}': Antigo='{existing_value}', Novo='{new_value}'")
            return True
    return False

def save_pdv_data(cursor, connection, batch_results):
    """Salva ou atualiza um lote de dados de PDVs no Oracle usando MERGE."""
    if not batch_results: 
        logger.info("Nenhum dado no batch para salvar no Oracle.")
        return 0, 0
        
    sql_merge = f"""
        MERGE INTO {TABLE_NAME} dest
        USING (
            SELECT :v_IP AS IP, :v_NROEMPRESA AS NROEMPRESA, :v_NROCHECKOUT AS NROCHECKOUT,
                   :v_SEGMENTO AS SEGMENTO, :v_OPERACAO AS OPERACAO, :v_PLACA_MAE AS PLACA_MAE,
                   :v_PROCESSADOR AS PROCESSADOR, :v_CORES_THREADS AS CORES_THREADS,
                   :v_RAM AS RAM, :v_DISCO AS DISCO, :v_ARMAZENAMENTO AS ARMAZENAMENTO,
                   :v_RELEASE AS RELEASE, :v_STATUS AS STATUS,
                   TO_DATE(:v_DTAINCLUSAO_STR, 'YYYY-MM-DD HH24:MI:SS') AS DTAINCLUSAO_VAL,
                   TO_DATE(:v_DTAATUALIZACAO_STR, 'YYYY-MM-DD HH24:MI:SS') AS DTAATUALIZACAO_VAL
            FROM dual
        ) src ON (dest.NROEMPRESA = src.NROEMPRESA AND dest.NROCHECKOUT = src.NROCHECKOUT)
        WHEN MATCHED THEN UPDATE SET
            dest.IP = src.IP, dest.SEGMENTO = src.SEGMENTO, dest.OPERACAO = src.OPERACAO,
            dest.PLACA_MAE = src.PLACA_MAE, dest.PROCESSADOR = src.PROCESSADOR,
            dest.CORES_THREADS = src.CORES_THREADS, dest.RAM = src.RAM, dest.DISCO = src.DISCO,
            dest.ARMAZENAMENTO = src.ARMAZENAMENTO, dest.RELEASE = src.RELEASE, dest.STATUS = src.STATUS,
            dest.DTAINCLUSAO = src.DTAINCLUSAO_VAL, dest.DTAATUALIZACAO = src.DTAATUALIZACAO_VAL
            WHERE DECODE(NVL(dest.IP, '#NULL#'), NVL(src.IP, '#NULL#'), 0, 1) = 1 OR
                  DECODE(NVL(dest.SEGMENTO, '#NULL#'), NVL(src.SEGMENTO, '#NULL#'), 0, 1) = 1 OR
                  DECODE(NVL(dest.OPERACAO, '#NULL#'), NVL(src.OPERACAO, '#NULL#'), 0, 1) = 1 OR
                  DECODE(NVL(dest.PLACA_MAE, '#NULL#'), NVL(src.PLACA_MAE, '#NULL#'), 0, 1) = 1 OR
                  DECODE(NVL(dest.PROCESSADOR, '#NULL#'), NVL(src.PROCESSADOR, '#NULL#'), 0, 1) = 1 OR
                  DECODE(NVL(dest.CORES_THREADS, '#NULL#'), NVL(src.CORES_THREADS, '#NULL#'), 0, 1) = 1 OR
                  DECODE(NVL(dest.RAM, '#NULL#'), NVL(src.RAM, '#NULL#'), 0, 1) = 1 OR
                  DECODE(NVL(dest.DISCO, '#NULL#'), NVL(src.DISCO, '#NULL#'), 0, 1) = 1 OR
                  DECODE(NVL(dest.ARMAZENAMENTO, '#NULL#'), NVL(src.ARMAZENAMENTO, '#NULL#'), 0, 1) = 1 OR
                  DECODE(NVL(dest.RELEASE, '#NULL#'), NVL(src.RELEASE, '#NULL#'), 0, 1) = 1 OR
                  DECODE(NVL(dest.STATUS, '#NULL#'), NVL(src.STATUS, '#NULL#'), 0, 1) = 1 OR
                  dest.DTAINCLUSAO <> src.DTAINCLUSAO_VAL OR
                  dest.DTAATUALIZACAO <> src.DTAATUALIZACAO_VAL
        WHEN NOT MATCHED THEN INSERT (IP, NROEMPRESA, NROCHECKOUT, SEGMENTO, OPERACAO, PLACA_MAE, PROCESSADOR, CORES_THREADS, RAM, DISCO, ARMAZENAMENTO, RELEASE, STATUS, DTAINCLUSAO, DTAATUALIZACAO)
            VALUES (src.IP, src.NROEMPRESA, src.NROCHECKOUT, src.SEGMENTO, src.OPERACAO, src.PLACA_MAE, src.PROCESSADOR, src.CORES_THREADS, src.RAM, src.DISCO, src.ARMAZENAMENTO, src.RELEASE, src.STATUS, src.DTAINCLUSAO_VAL, src.DTAATUALIZACAO_VAL)
    """
    try:
        cursor.executemany(sql_merge, batch_results, batcherrors=True)
        connection.commit()
        rows_affected_count = cursor.rowcount
        batch_size = len(batch_results)
        for error in cursor.getbatcherrors(): 
            logger.error(f"Erro no batch Oracle na linha {error.offset}: {error.message}")
        logger.info(f"Lote de {batch_size} registros enviado ao Oracle (afetados: {rows_affected_count}).")
        return rows_affected_count, batch_size
    except Exception as e:
        logger.error(f"Erro crítico ao salvar dados no Oracle: {str(e)}", exc_info=True)
        tqdm.write(f"{RED}Erro DB/desconhecido ao salvar no Oracle: {str(e)}")
        if isinstance(e, oracledb.DatabaseError): 
            try: connection.rollback(); logger.info("Rollback da transação Oracle realizado.")
            except Exception as rb_e: logger.error(f"Erro ao tentar rollback: {rb_e}")
        return 0, len(batch_results) 

def mark_inactive_pdvs(cursor, connection, df_pdvs_ativos):
    """Marca PDVs como 'Inativo' se não estiverem mais na lista de PDVs ativos."""
    try:
        cursor.execute(f"SELECT NROEMPRESA, NROCHECKOUT FROM {TABLE_NAME} WHERE STATUS <> 'Inativo'")
        existing_pdvs_in_table = {(int(row[0]), int(row[1])) for row in cursor.fetchall()}
        
        active_pdvs_from_query = set()
        if not df_pdvs_ativos.empty:
            active_pdvs_from_query = {(int(row["NROEMPRESA"]), int(row["NROCHECKOUT"])) for _, row in df_pdvs_ativos.iterrows()}
        
        pdvs_to_mark_inactive = existing_pdvs_in_table - active_pdvs_from_query
        
        if pdvs_to_mark_inactive:
            current_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            update_payload = [{'new_status': 'Inativo', 'dtaatualizacao': current_time_str, 'nroempresa': ne, 'nrocheckout': nc} for ne, nc in pdvs_to_mark_inactive]
            
            cursor.executemany(
                f"UPDATE {TABLE_NAME} SET STATUS = :new_status, DTAATUALIZACAO = TO_DATE(:dtaatualizacao, 'YYYY-MM-DD HH24:MI:SS') WHERE NROEMPRESA = :nroempresa AND NROCHECKOUT = :nrocheckout",
                update_payload
            )
            connection.commit()
            logger.info(f"{len(pdvs_to_mark_inactive)} PDVs candidatos a 'Inativo'. Afetados no DB: {cursor.rowcount}.")
            # tqdm.write(f"{YELLOW}{cursor.rowcount} de {len(pdvs_to_mark_inactive)} PDVs marcados como 'Inativo'.") # Removido do terminal
        else:
<<<<<<< HEAD
            logger.info("Nenhum PDV encontrado para marcar como 'Inativo' nesta execução.")
            # tqdm.write(f"{GREEN}Nenhum PDV para marcar como 'Inativo'.") # Removido do terminal
            
=======
            return ram_quantidade
    else:
        return "Não detectado"

def obter_info_disco(cliente):
    """
    Obtém informações sobre o disco (tipo e capacidade).
    """
    # Primeiro, identificar discos principais
    cmd_listar_discos = "lsblk -d | grep -v loop | grep -v sr | awk '{print $1}' | head -3"
    resultado_discos = cliente.execute_command(cmd_listar_discos)
    discos = resultado_discos.strip().split('\n') if resultado_discos else ["sda"]

    # Tipo de disco - Tentando em múltiplos discos possíveis
    tipo_disco = "Não detectado"
    for disco in discos:
        # Corrigindo a sintaxe do lsblk
        comandos_tipo_disco = [
            f"lsblk -d -o name,rota /dev/{disco} 2>/dev/null | grep {disco} | awk '{{print $2}}'",
            f"cat /sys/block/{disco}/queue/rotational 2>/dev/null",
            f"smartctl -a /dev/{disco} 2>/dev/null | grep -i 'rotation rate\\|solid state device'",
            f"lshw -class disk | grep -i 'logical name.*{disco}' -A5 | grep -i 'solid\\|ssd\\|rotation'",
            # Método alternativo com hdparm
            f"hdparm -I /dev/{disco} 2>/dev/null | grep -i 'nominal media rotation rate\\|solid state'",
            # Verificar nome do modelo que frequentemente indica SSD
            f"smartctl -i /dev/{disco} 2>/dev/null | grep -i 'device model' | grep -i 'ssd'"
        ]
        for comando in comandos_tipo_disco:
            resultado = cliente.execute_command(comando)
            if resultado:
                resultado = resultado.lower()
                if "0" in resultado or "solid" in resultado or "ssd" in resultado or "not a rotating device" in resultado:
                    tipo_disco = "SSD"
                    break
                elif "1" in resultado or "rpm" in resultado or "rotational" in resultado:
                    tipo_disco = "HD"
                    break
        if tipo_disco != "Não detectado":
            break

    # Método alternativo baseado na velocidade de leitura (SSDs são mais rápidos)
    if tipo_disco == "Não detectado":
        for disco in discos:
            cmd_velocidade = f"hdparm -t /dev/{disco} 2>/dev/null | grep 'Timing buffered' || echo ''"
            resultado = cliente.execute_command(cmd_velocidade)
            if resultado and "MB/sec" in resultado:
                try:
                    # Extrair a velocidade
                    search_result = re.search(r'(\d+\.\d+) MB/sec', resultado)
                    if search_result is not None:
                        velocidade = float(search_result.group(1))
                        # SSDs tipicamente têm velocidades maiores que 100 MB/sec
                        if velocidade > 100:
                            tipo_disco = "SSD"
                        else:
                            tipo_disco = "HD"
                        break
                except:
                    pass

    # Último recurso: inferir pelo nome do dispositivo ou tamanho
    if tipo_disco == "Não detectado":
        cmd_info_discos = "lsblk -d -b"
        resultado = cliente.execute_command(cmd_info_discos)
        if resultado:
            # SSDs comerciais pequenos comuns (até 256GB), HDs raramente abaixo de 320GB
            for linha in resultado.splitlines():
                if "loop" in linha or "sr" in linha:
                    continue
                partes = linha.split()
                if len(partes) >= 4:
                    try:
                        tamanho = int(partes[3])
                        if "nvme" in partes[0]:  # NVMe são sempre SSDs
                            tipo_disco = "SSD"
                            break
                        # Inferência por tamanho (SSDs pequenos são comuns)
                        if tamanho <= 256_000_000_000:  # ~256GB
                            tipo_disco = "SSD"
                            break
                        if tamanho >= 1_000_000_000_000:  # ~1TB
                            tipo_disco = "HD"  # Mais provável ser HD para discos grandes
                    except:
                        pass

    # Capacidade de armazenamento - Comando melhorado para obter o tamanho total do disco físico
    comandos_armazenamento = [
        # Comandos específicos para tamanho físico do disco
        "lsblk -dbn -o SIZE /dev/sda 2>/dev/null || echo 'Não identificado'",
        "fdisk -l 2>/dev/null | grep 'Disk /dev/sda' | awk '{print $5}'",
        "smartctl -i /dev/sda 2>/dev/null | grep 'User Capacity' | cut -d '[' -f2 | cut -d ']' -f1",
        # Comandos de backup
        "lsblk -b -d -o size | head -2 | tail -1",
        "df -B1 --total | grep total | awk '{print $2}'"
    ]
    armazenamento = "Não detectado"
    for comando in comandos_armazenamento:
        resultado = cliente.execute_command(comando)
        if resultado:
            try:
                # Se o resultado for numérico (bytes)
                if resultado.isdigit():
                    tamanho_bytes = int(resultado)
                    if tamanho_bytes > 0:  # Ignorar valores zero
                        if tamanho_bytes >= 1_000_000_000_000:  # 1TB ou mais
                            armazenamento = f"{tamanho_bytes / 1_000_000_000_000:.1f}TB"
                        else:
                            armazenamento = f"{tamanho_bytes / 1_000_000_000:.0f}GB"
                        break
                else:
                    # Procurar padrões como "120GB" ou "500 GB" ou "1TB"
                    match = re.search(r"(\d+\.?\d*)([GMKTP]i?B?)", resultado)
                    if match is not None:
                        valor = float(match.group(1))
                        unidade = match.group(2)[0].upper()  # G, T, etc.
                        if unidade == "T":
                            armazenamento = f"{valor}TB"
                        elif unidade == "G":
                            armazenamento = f"{valor}GB"
                        elif unidade == "M":
                            armazenamento = f"{valor / 1000:.1f}GB"
                        break
            except Exception as e:
                logging.error(f"Erro ao processar tamanho do disco: {str(e)}")
                continue

    # Se não conseguiu detectar o tamanho, tenta um último método mais agressivo
    if armazenamento == "Não detectado" or armazenamento == "0GB":
        try:
            # Listar todos os dispositivos de bloco e somar seus tamanhos
            cmd = "lsblk -dbn -o NAME,SIZE | grep -v loop | grep -v sr | awk '{sum += $2} END {print sum}'"
            resultado = cliente.execute_command(cmd)
            if resultado and resultado.isdigit() and int(resultado) > 0:
                tamanho_bytes = int(resultado)
                if tamanho_bytes >= 1_000_000_000_000:  # 1TB ou mais
                    armazenamento = f"{tamanho_bytes / 1_000_000_000_000:.1f}TB"
                else:
                    armazenamento = f"{tamanho_bytes / 1_000_000_000:.0f}GB"
        except Exception as e:
            logging.error(f"Erro no método alternativo de detecção de disco: {str(e)}")
    
    return {"tipo": tipo_disco, "capacidade": armazenamento}

def get_hardware_info(ip, username, password, retry=0):
    """
    Coleta todas as informações de hardware de um único PDV
    """
    try:
        # Criar cliente SSH e conectar
        client = SSHClient(ip, username, password)
        if not client.connect():
            return {"erro": "Falha na conexão SSH"}

        # Coletar informações usando métodos da versão original
        fabricante_placa_mae = obter_fabricante_placa_mae(client)
        modelo_placa_mae = obter_modelo_placa_mae(client)
        info_processador = obter_info_processador(client)
        info_ram = obter_info_ram(client)
        info_disco = obter_info_disco(client)
        iso_release = obter_iso_release(client)

        # Fechar conexão
        client.close()

        # Combinar fabricante e modelo da placa-mãe
        placa_mae = "Não detectado"
        if fabricante_placa_mae != "Não detectado" and modelo_placa_mae != "Não detectado":
            placa_mae = f"{fabricante_placa_mae} - {modelo_placa_mae}"
        elif fabricante_placa_mae != "Não detectado":
            placa_mae = fabricante_placa_mae
        elif modelo_placa_mae != "Não detectado":
            placa_mae = modelo_placa_mae

        return {
            "placa-mãe": placa_mae,
            "processador": info_processador["modelo"],
            "cores/threads": info_processador["nucleos_threads"],
            "ram": info_ram,
            "disco": info_disco["tipo"],
            "armazenamento": info_disco["capacidade"],
            "release": iso_release
        }
>>>>>>> bb3374c8ebbc2839fd2ae95848523d2e15acb503
    except Exception as e:
        logger.error(f"Erro ao marcar PDVs inativos: {str(e)}", exc_info=True)
        # tqdm.write(f"{RED}Erro ao marcar PDVs inativos: {str(e)}") # Opcional: remover ou manter para erros no terminal

<<<<<<< HEAD
def save_to_csv(batch_results_raw):
    """Salva um lote de dados de PDVs em um arquivo CSV."""
    if not batch_results_raw: return 0
    
    batch_results_for_csv = []
    for record_raw in batch_results_raw:
        record_csv = {key.replace('v_', ''): value for key, value in record_raw.items()}
        if 'DTAINCLUSAO_STR' in record_csv: record_csv['DTAINCLUSAO'] = record_csv.pop('DTAINCLUSAO_STR')
        if 'DTAATUALIZACAO_STR' in record_csv: record_csv['DTAATUALIZACAO'] = record_csv.pop('DTAATUALIZACAO_STR')
        batch_results_for_csv.append(record_csv)
        
    try:
        fieldnames = ["IP", "NROEMPRESA", "NROCHECKOUT", "SEGMENTO", "OPERACAO", "PLACA_MAE", 
                      "PROCESSADOR", "CORES_THREADS", "RAM", "DISCO", "ARMAZENAMENTO", 
                      "RELEASE", "STATUS", "DTAINCLUSAO", "DTAATUALIZACAO"]
        file_exists = os.path.exists(CSV_FILE)
        with open(CSV_FILE, mode='a' if file_exists else 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction='ignore')
            if not file_exists: writer.writeheader()
            writer.writerows(batch_results_for_csv)
        logger.info(f"Salvos {len(batch_results_for_csv)} registros no CSV {CSV_FILE}.")
        return len(batch_results_for_csv)
    except Exception as e:
        logger.error(f"Erro ao salvar dados no CSV {CSV_FILE}: {str(e)}", exc_info=True)
        tqdm.write(f"{RED}Erro ao salvar dados no CSV: {str(e)}")
        return 0
=======
def extrair_info_pdv(descricao):
    """
    Extrai o número do PDV a partir da string de descrição.
    """
    if not isinstance(descricao, str):
        return "Desconhecido"
    # Procurar por padrões como "PDV X" ou "PDV-X" na descrição
    match = re.search(r'PDV[\s-]*(\d+)', descricao)
    if match is not None:
        return f"PDV {match.group(1)}"
    # Se não encontrar o padrão PDV, verificar por SelfCheckout
    if "SelfCheckout" in descricao:
        return "SelfCheckout"
    # Caso não encontre nada específico
    return descricao
>>>>>>> bb3374c8ebbc2839fd2ae95848523d2e15acb503

def handle_table_not_found(cursor):
    """Interage com o usuário quando a tabela de hardware não é encontrada."""
    print(f"{YELLOW}Tabela {TABLE_NAME} não encontrada no banco de dados.")
    while True:
        print(f"{CYAN}Escolha uma opção:\n  1. Criar tabela {TABLE_NAME}\n  2. Exportar dados apenas para CSV ({CSV_FILE})\n  3. Sair")
        choice = input(f"{CYAN}Digite 1, 2 ou 3: {RESET}").strip()
        if choice == '1':
            if create_oracle_table(cursor): return "ORACLE"
            else: print(f"{RED}Falha ao criar tabela. Verifique logs e permissões. Tente novamente ou escolha outra opção.")
        elif choice == '2':
            print(f"{GREEN}Dados coletados serão salvos apenas em {CSV_FILE}."); return "CSV"
        elif choice == '3':
            print(f"{MAGENTA}Script encerrado pelo usuário."); exit(0)
        else: print(f"{YELLOW}Opção inválida. Por favor, digite 1, 2 ou 3.")

<<<<<<< HEAD
# --- Lógica Principal ---
=======
# Lista de processadores obsoletos
PROCESSADORES_OBSOLETOS = [
    "Intel(R) Celeron(R) CPU G1820",
    "Intel(R) Celeron(R) CPU J1800",
    "Intel(R) Celeron(R) CPU G3900",
    "Intel(R) Atom(TM) CPU D2550",
    "Intel(R) Atom(TM) CPU D2500",
    "Intel(R) Celeron(R) CPU G3930",
    "Intel(R) Pentium(R) CPU G620",
    "Pentium(R) Dual-Core CPU E5300",
    "Pentium(R) Dual-Core CPU E5800",
    "Intel(R) Celeron(R) CPU 847",
    "Intel(R) Core(TM) i5-2400 CPU @ 3.10GHz"
]

# Função para verificar se o processador é obsoleto
def is_processador_obsoleto(processador):
    if pd.isna(processador) or str(processador).strip() == "":
        return False
    return any(proc in str(processador) for proc in PROCESSADORES_OBSOLETOS)

# Função para verificar se a RAM está abaixo do ideal (menos de 4GB)
def is_ram_baixa(ram):
    if pd.isna(ram) or str(ram).strip() == "":
        return False
    try:
        ram_str = str(ram)
        if "GB" not in ram_str:
            return False
        # Extrai apenas os números antes de "GB"
        capacidade_str = ''.join(c for c in ram_str.split("GB")[0] if c.isdigit())
        if not capacidade_str:
            return False
        capacidade_gb = int(capacidade_str)
        return capacidade_gb < 4
    except (AttributeError, ValueError):
        return False

# Função para verificar se o disco é HD
def is_hd_lento(disco):
    if pd.isna(disco):
        return False
    return "HD" in str(disco)

# Função para determinar a prioridade
def determinar_prioridade(row):
    if pd.isna(row["PROCESSADOR"]) or str(row["PROCESSADOR"]).strip() == "":
        return "Offline"
    
    processador_obsoleto = is_processador_obsoleto(row["PROCESSADOR"])
    ram_baixa = is_ram_baixa(row["RAM"])
    hd_lento = is_hd_lento(row["DISCO"])

    if processador_obsoleto:
        return "Alta"
    elif ram_baixa and hd_lento:
        return "Alta"
    elif ram_baixa or hd_lento:
        return "Média"
    else:
        return ""  # Prioridade baixa fica em branco

# Função para determinar o motivo do upgrade
def determinar_motivo(row):
    motivos = []

    if pd.isna(row["PROCESSADOR"]) or str(row["PROCESSADOR"]).strip() == "":
        return ""
    
    if is_processador_obsoleto(row["PROCESSADOR"]):
        motivos.append("Processador em obsolência")
    if is_ram_baixa(row["RAM"]):
        motivos.append("RAM abaixo do ideal")
    if is_hd_lento(row["DISCO"]):
        motivos.append("HD lento")

    return "; ".join(motivos) if motivos else ""

>>>>>>> bb3374c8ebbc2839fd2ae95848523d2e15acb503
def main():
    """Função principal para orquestrar a coleta e salvamento de dados de hardware dos PDVs."""
    try:
        from coletaPDV import get_hardware_info 
    except ImportError:
        logger.critical("ERRO FATAL: coletaPDV.py não encontrado ou função get_hardware_info não definida.")
        print(f"{RED}ERRO FATAL: coletaPDV.py não encontrado ou get_hardware_info não definida nele. Verifique o arquivo.")
        exit(1)

    db_connection = None; db_cursor = None; batch_buffer = []; processed_pdvs_count = 0
    output_mode = "ORACLE"; sucessos_coleta_efetiva = 0; total_registros_enviados_db = 0
    total_registros_afetados_db = 0; total_registros_salvos_csv = 0

    print(f"{MAGENTA}=== Iniciando Coleta de Hardware de PDVs ===")
    logger.info("=== INÍCIO DA EXECUÇÃO DO SCRIPT DE COLETA DE HARDWARE ===")

    try:
        configs = load_env_configs()
        
        print(f"{CYAN}Conectando ao Oracle ({configs['DSN']})...")
        db_connection = oracledb.connect(user=configs["ORACLE_USER"], password=configs["ORACLE_PASSWORD"], dsn=configs["DSN"])
        db_cursor = db_connection.cursor()
        print(f"{GREEN}Conexão Oracle estabelecida.")

        table_status = check_table_exists(db_cursor)
        if table_status == "TABLE_EXISTS": print(f"{GREEN}Tabela {TABLE_NAME} encontrada e pronta para uso.")
        elif table_status == "TABLE_NOT_FOUND": output_mode = handle_table_not_found(db_cursor)
        else: print(f"{RED}Erro crítico ao verificar tabela {TABLE_NAME}. Verifique os logs. Saindo."); return

        print(f"{CYAN}Buscando PDVs ativos da fonte de dados...")
        query_pdvs_ativos = """
            WITH PARAMPDV AS (
                SELECT NROEMPRESA, NROCHECKOUT, COALESCE(TO_CHAR(VALOR), 'VendaCupomNota') AS VALOR
                FROM CONSINCOMONITOR.TB_PARAMPDVVALOR WHERE PARAMPDV = 'ModoOperacao'
            ), PDV_ATIVO AS (
                SELECT IP, NROEMPRESA, NROCHECKOUT, NROSEGMENTO FROM CONSINCOMONITOR.TB_CHECKOUT
                WHERE ATIVO = 'S' AND NROCHECKOUT <> 100
            )
            SELECT A.IP, A.NROEMPRESA, A.NROCHECKOUT, S.SEGMENTO, COALESCE(B.VALOR, 'PDV') AS OPERACAO
            FROM PDV_ATIVO A
            LEFT JOIN PARAMPDV B ON A.NROEMPRESA = B.NROEMPRESA AND A.NROCHECKOUT = B.NROCHECKOUT
            JOIN CONSINCOMONITOR.TB_SEGMENTO S ON S.NROSEGMENTO = A.NROSEGMENTO
            ORDER BY A.NROEMPRESA, A.NROCHECKOUT
        """
        db_cursor.execute(query_pdvs_ativos)
        pdv_data_from_db = db_cursor.fetchall()
        df_pdvs = pd.DataFrame(pdv_data_from_db, columns=[d[0].upper() for d in db_cursor.description]) if pdv_data_from_db else pd.DataFrame()

        if df_pdvs.empty:
            logger.warning("Nenhum PDV ativo encontrado na consulta à TB_CHECKOUT.")
            print(f"{YELLOW}Nenhum PDV ativo encontrado para processar.")
            if output_mode == "ORACLE":
                logger.info("Verificando PDVs para marcar como 'Inativo' (nenhum PDV ativo encontrado).")
                mark_inactive_pdvs(db_cursor, db_connection, df_pdvs)
            print(f"{CYAN}Processo concluído."); return

        df_pdvs["NROEMPRESA"] = pd.to_numeric(df_pdvs["NROEMPRESA"])
        df_pdvs["NROCHECKOUT"] = pd.to_numeric(df_pdvs["NROCHECKOUT"])
        logger.info(f"Encontrados {len(df_pdvs)} PDVs ativos para processar.")
        print(f"{GREEN}Encontrados {len(df_pdvs)} PDVs ativos.")

        pdvs_args_list = [
            ({"IP": r["IP"], "NROEMPRESA": int(r["NROEMPRESA"]), 
              "NROCHECKOUT": int(r["NROCHECKOUT"]), "SEGMENTO": r["SEGMENTO"], 
              "OPERACAO": r["OPERACAO"]}, 
             configs["SSH_USERNAME"], configs["SSH_PASSWORD"]) 
            for _, r in df_pdvs.iterrows()
        ]
        total_pdvs_to_scan = len(pdvs_args_list)
        num_workers = min(configs["MAX_WORKERS"], total_pdvs_to_scan if total_pdvs_to_scan > 0 else 1)
        print(f"{CYAN}Iniciando coleta de {total_pdvs_to_scan} PDVs com {num_workers} workers...")
        
<<<<<<< HEAD
        with tqdm(total=total_pdvs_to_scan, desc=f"{CYAN}Escaneando PDVs", 
                  bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]",
                  ncols=100) as progress_bar:
            with concurrent.futures.ThreadPoolExecutor(max_workers=num_workers) as executor:
                future_to_pdv_map = {
                    executor.submit(get_hardware_info, pa[0]["IP"], pa[1], pa[2]): pa[0] 
                    for pa in pdvs_args_list
                }

                for future in concurrent.futures.as_completed(future_to_pdv_map):
                    pdv_initial_info = future_to_pdv_map[future]
                    ip = pdv_initial_info["IP"]; nroempresa = pdv_initial_info["NROEMPRESA"]; nrocheckout = pdv_initial_info["NROCHECKOUT"]
                    current_time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                    data_for_db = {
                        "IP": ip, "NROEMPRESA": nroempresa, "NROCHECKOUT": nrocheckout,
                        "SEGMENTO": pdv_initial_info["SEGMENTO"], "OPERACAO": pdv_initial_info["OPERACAO"]
                    }
                    for fld in HARDWARE_FIELDS: data_for_db[fld] = "Não detectado"

                    coleta_online = False 
                    hw_collection_result = None
                    status_retornado_coleta_debug = "Coleta não iniciada"

=======
        # Solicitar credenciais
        username = input("Digite o nome de usuário para SSH (padrão: root): ") or "root"
        password = input("Digite a senha para SSH: ")
        
        # Criar lista de IPs com informações PDV para processamento
        ips_to_process = []
        for indice, linha in df_pdvs.iterrows():
            ip = linha["IP"]
            # Obter informação do PDV a partir das novas colunas
            if "DESCRICAO" in df_pdvs.columns:
                info_pdv = extrair_info_pdv(linha["DESCRICAO"])
            else:
                info_pdv = f"IP {ip}"  # Fallback se não existir coluna DESCRICAO
            ips_to_process.append((ip, username, password, info_pdv))
        
        # Número de threads a usar (ajustar conforme hardware disponível)
        num_workers = min(MAX_WORKERS, len(ips_to_process))
        print(f"Processando {len(ips_to_process)} IPs com {num_workers} threads paralelas...")
        
        # Dicionário para armazenar resultados
        results = {}
        progress_bar = tqdm(total=len(ips_to_process), desc="Escaneando PDVs")
        
        # Processar IPs em paralelo
        with concurrent.futures.ThreadPoolExecutor(max_workers=num_workers) as executor:
            future_to_ip = {executor.submit(process_ip, args): args for args in ips_to_process}
            # Salvar progresso parcial a cada N IPs processados
            save_interval = max(5, min(20, len(ips_to_process) // 10))  # Ajuste dinâmico
            completed = 0
            for future in concurrent.futures.as_completed(future_to_ip):
                ip, info = future.result()
                results[ip] = info
                # Atualizar barra de progresso
                progress_bar.update(1)
                completed += 1
                # Salvar progresso parcial
                if completed % save_interval == 0 or completed == len(ips_to_process):
                    # Atualizar DataFrame com resultados obtidos até agora, aplicando regras de validação
                    for idx, row in df_pdvs.iterrows():
                        ip = row["IP"]
                        if ip in results:
                            for key, value in results[ip].items():
                                column_key = key.upper()
                                # Determinar a coluna correta para atualização
                                target_column = None
                                if column_key in mapeamento_colunas and mapeamento_colunas[column_key]:
                                    # Caso seja uma coluna que foi renomeada
                                    target_column = mapeamento_colunas[column_key]
                                elif column_key in df_pdvs.columns:
                                    # Caso seja uma coluna que manteve o nome
                                    target_column = column_key
                                
                                # Aplicar regras de validação se a coluna existe
                                if target_column:
                                    current_value = df_pdvs.at[idx, target_column]
                                    # Regra 1: Se a célula estiver vazia, preencher com os novos dados
                                    if pd.isna(current_value) or current_value == "" or current_value == "Não detectado":
                                        df_pdvs.at[idx, target_column] = value
                                    # Regra 2: Se os valores forem diferentes, substituir os dados existentes
                                    elif str(current_value).strip() != str(value).strip():
                                        df_pdvs.at[idx, target_column] = value
                                    # Se os valores forem iguais, manter os dados existentes (não faz nada)
                    # Salvar arquivo temporário
>>>>>>> bb3374c8ebbc2839fd2ae95848523d2e15acb503
                    try:
                        hw_collection_result = future.result() 
                        logger.info(f"PDV {nroempresa}-{nrocheckout} (IP: {ip}): Resultado bruto da coleta: {hw_collection_result}")
                        
                        status_retornado_coleta_debug = hw_collection_result.get("status") # Chave "status" do coletaPDV.py
                        logger.info(f"PDV {nroempresa}-{nrocheckout} (IP: {ip}): 'status' de coletaPDV.py: '{status_retornado_coleta_debug}'")

                        if status_retornado_coleta_debug == "SUCESSO" or status_retornado_coleta_debug == "FALHA COMANDOS":
                            coleta_online = True
                            if status_retornado_coleta_debug == "FALHA COMANDOS":
                                logger.warning(f"PDV {nroempresa}-{nrocheckout} (IP: {ip}): Coleta retornou '{status_retornado_coleta_debug}'. PDV será ONLINE, hardware pode estar incompleto. Raw: {hw_collection_result}")
                    
                    except Exception as e_collect_thread:
                        logger.error(f"PDV {nroempresa}-{nrocheckout} (IP: {ip}): EXCEÇÃO na thread de coleta: {type(e_collect_thread).__name__} - {str(e_collect_thread)}", exc_info=True)
                        tqdm.write(f"{RED}Exceção na coleta do PDV {nroempresa}-{nrocheckout} (IP: {ip}): {type(e_collect_thread).__name__}")
                        status_retornado_coleta_debug = f"Exceção na thread: {type(e_collect_thread).__name__}"
                    
                    if coleta_online and hw_collection_result:
                        data_for_db["STATUS"] = "ONLINE"
                        for fld_key_db in HARDWARE_FIELDS: 
                            data_for_db[fld_key_db] = hw_collection_result.get(fld_key_db.lower(), "Não detectado")
                    else:
                        data_for_db["STATUS"] = "OFFLINE"
                        logger.warning(f"PDV {nroempresa}-{nrocheckout} (IP: {ip}) MARCADO COMO OFFLINE. Causa/Status da coleta: '{status_retornado_coleta_debug}'. Resultado bruto (se houver): {hw_collection_result}")
                    
                    existing_data_from_db = get_existing_data(db_cursor, nroempresa, nrocheckout) if output_mode == "ORACLE" else None
                    needs_db_write = False
                    dados_realmente_mudaram = False

                    if existing_data_from_db:
                        if data_for_db["STATUS"] == "OFFLINE":
                            for fld in HARDWARE_FIELDS: 
                                data_for_db[fld] = existing_data_from_db.get(fld, "Não detectado")
                        
                        dados_realmente_mudaram = has_relevant_data_changes(data_for_db, existing_data_from_db)

                        if data_for_db["STATUS"] == "ONLINE":
                            data_for_db["DTAATUALIZACAO"] = current_time_str
                            if dados_realmente_mudaram:
                                data_for_db["DTAINCLUSAO"] = current_time_str 
                            else: 
                                data_for_db["DTAINCLUSAO"] = existing_data_from_db["DTAINCLUSAO"].strftime('%Y-%m-%d %H:%M:%S')
                        else: # STATUS == "OFFLINE"
                            data_for_db["DTAINCLUSAO"] = existing_data_from_db["DTAINCLUSAO"].strftime('%Y-%m-%d %H:%M:%S')
                            data_for_db["DTAATUALIZACAO"] = current_time_str
                        
                        needs_db_write = True 
                    
                    else: # Novo PDV
                        data_for_db["DTAINCLUSAO"] = current_time_str
                        data_for_db["DTAATUALIZACAO"] = current_time_str
                        needs_db_write = True
                        dados_realmente_mudaram = True 
                    
                    if needs_db_write:
                        batch_entry = {
                            'v_IP': data_for_db.get("IP"), 'v_NROEMPRESA': data_for_db.get("NROEMPRESA"),
                            'v_NROCHECKOUT': data_for_db.get("NROCHECKOUT"), 'v_SEGMENTO': data_for_db.get("SEGMENTO"),
                            'v_OPERACAO': data_for_db.get("OPERACAO"),
                            'v_PLACA_MAE': data_for_db.get("PLACA_MAE", "Não detectado"),
                            'v_PROCESSADOR': data_for_db.get("PROCESSADOR", "Não detectado"),
                            'v_CORES_THREADS': data_for_db.get("CORES_THREADS", "Não detectado"),
                            'v_RAM': data_for_db.get("RAM", "Não detectado"),
                            'v_DISCO': data_for_db.get("DISCO", "Não detectado"),
                            'v_ARMAZENAMENTO': data_for_db.get("ARMAZENAMENTO", "Não detectado"),
                            'v_RELEASE': data_for_db.get("RELEASE", "Não detectado"),
                            'v_STATUS': data_for_db.get("STATUS"),
                            'v_DTAINCLUSAO_STR': data_for_db.get("DTAINCLUSAO"), 
                            'v_DTAATUALIZACAO_STR': data_for_db.get("DTAATUALIZACAO") 
                        }
                        batch_buffer.append(batch_entry)
                        
                        if data_for_db["STATUS"] == "ONLINE":
                            if not existing_data_from_db or dados_realmente_mudaram or \
                               (existing_data_from_db and existing_data_from_db.get("STATUS") != "ONLINE"):
                                sucessos_coleta_efetiva += 1

                        if len(batch_buffer) >= configs["SAVE_INTERVAL"]:
                            if output_mode == "ORACLE":
                                afetados, enviados = save_pdv_data(db_cursor, db_connection, batch_buffer)
                                total_registros_afetados_db += afetados; total_registros_enviados_db += enviados
                            else: total_registros_salvos_csv += save_to_csv(batch_buffer)
                            batch_buffer = []
                    
                    processed_pdvs_count +=1
                    progress_bar.update(1)
                    progress_bar.set_postfix_str(f"Último: {ip} ({data_for_db.get('STATUS', 'N/A')})")

        if batch_buffer:
            if output_mode == "ORACLE":
                afetados, enviados = save_pdv_data(db_cursor, db_connection, batch_buffer)
                total_registros_afetados_db += afetados; total_registros_enviados_db += enviados
            else: total_registros_salvos_csv += save_to_csv(batch_buffer)

        if output_mode == "ORACLE": 
            logger.info("Verificando PDVs para marcar como 'Inativo'...")
            mark_inactive_pdvs(db_cursor, db_connection, df_pdvs)

        print(f"\n{GREEN}=== Processo de Coleta Concluído ===")
        print(f"{CYAN}Total de PDVs na consulta de ativos: {total_pdvs_to_scan}")
        print(f"{CYAN}PDVs processados nesta execução: {processed_pdvs_count}")
        print(f"{GREEN}PDVs 'ONLINE' com escrita significativa: {sucessos_coleta_efetiva}")
        pdvs_outros_resultados = processed_pdvs_count - sucessos_coleta_efetiva
        if pdvs_outros_resultados >= 0: 
            print(f"{YELLOW}PDVs 'OFFLINE' ou 'ONLINE' sem mudanças de dados que contassem como 'sucesso': {pdvs_outros_resultados}")
        
        if output_mode == "ORACLE":
            print(f"{CYAN}Registros enviados para processamento no Oracle: {total_registros_enviados_db}")
            print(f"{CYAN}Registros efetivamente inseridos/atualizados no Oracle: {total_registros_afetados_db}")
        elif output_mode == "CSV": 
            print(f"{CYAN}Registros salvos no arquivo CSV: {total_registros_salvos_csv}")

    except oracledb.DatabaseError as oe: 
        error_obj, = oe.args
        logger.critical(f"Erro crítico de banco de dados Oracle: {str(oe)} (Código: {error_obj.code})", exc_info=True)
        print(f"{RED}Erro crítico de banco de dados Oracle: {str(oe)}")
        print(f"{YELLOW}Verifique credenciais, DSN, listener e permissões.")
    except ImportError as imp_err: 
        logger.critical(f"Erro de importação não tratado anteriormente: {str(imp_err)}", exc_info=True)
        print(f"{RED}Erro de Importação Crítico: {str(imp_err)}")
    except Exception as e_main: 
        logger.critical(f"Erro INESPERADO e crítico na execução principal: {str(e_main)}", exc_info=True)
        print(f"{RED}Erro Crítico Inesperado na Aplicação: {str(e_main)}")
    finally:
        if db_cursor:
            try: db_cursor.close(); print(f"{CYAN}Cursor Oracle fechado.")
            except Exception as e: logger.warning(f"Erro ao fechar cursor Oracle: {str(e)}", exc_info=True)
        if db_connection:
            try: db_connection.close(); print(f"{CYAN}Conexão Oracle fechada.")
            except Exception as e: logger.warning(f"Erro ao fechar conexão Oracle: {str(e)}", exc_info=True)
        
<<<<<<< HEAD
        print(f"{CYAN}Detalhes completos da execução no arquivo de log: {YELLOW}{LOG_FILE}{CYAN}.")
        if output_mode == "CSV": 
             print(f"{CYAN}Os dados (se houver) foram exportados para: {YELLOW}{CSV_FILE}{CYAN}.")
        logger.info("=== FIM DA EXECUÇÃO DO SCRIPT DE COLETA DE HARDWARE ===")
=======
        # Atualizar DataFrame com todos os resultados, aplicando regras de validação
        for idx, row in df_pdvs.iterrows():
            ip = row["IP"]
            if ip in results:
                for key, value in results[ip].items():
                    column_key = key.upper()
                    # Verificar se existe mapeamento para essa coluna
                    mapped_column = None
                    for old_col, new_col in mapeamento_colunas.items():
                        if old_col.lower() == key.lower():
                            mapped_column = new_col
                            break
                    
                    # Determinar a coluna correta para atualização
                    target_column = None
                    if mapped_column and mapped_column in df_pdvs.columns:
                        target_column = mapped_column
                    elif column_key in df_pdvs.columns:
                        target_column = column_key
                    
                    # Aplicar regras de validação se a coluna existe
                    if target_column:
                        current_value = df_pdvs.at[idx, target_column]
                        # Regra 1: Se a célula estiver vazia, preencher com os novos dados
                        if pd.isna(current_value) or current_value == "" or current_value == "Não detectado":
                            df_pdvs.at[idx, target_column] = value
                            logging.info(f"IP {ip}, coluna {target_column}: Célula vazia preenchida com '{value}'")
                        # Regra 2: Se os valores forem diferentes, substituir os dados existentes
                        elif str(current_value).strip() != str(value).strip():
                            old_value = current_value
                            df_pdvs.at[idx, target_column] = value
                            logging.info(f"IP {ip}, coluna {target_column}: Valor atualizado de '{old_value}' para '{value}'")
                        # Se os valores forem iguais, manter os dados existentes (não faz nada)
        
        # Remover a coluna MODELO_PLACA_MAE já que foi combinada com FABRICANTE_PLACA_MAE
        if "MODELO_PLACA_MAE" in df_pdvs.columns:
            df_pdvs = df_pdvs.drop(columns=["MODELO_PLACA_MAE"])
        
        # Adicionar colunas de prioridade e motivo do upgrade
        if "PRIORIDADE" not in df_pdvs.columns:
            df_pdvs["PRIORIDADE"] = ""
        if "MOTIVO DO UPGRADE" not in df_pdvs.columns:
            df_pdvs["MOTIVO DO UPGRADE"] = ""
        
        # Aplicar funções para determinar prioridade e motivo
        print("Analisando prioridades de upgrade...")
        df_pdvs["PRIORIDADE"] = df_pdvs.apply(determinar_prioridade, axis=1)
        df_pdvs["MOTIVO DO UPGRADE"] = df_pdvs.apply(determinar_motivo, axis=1)
        
        # Salvar resultados finais com formatação de cores
        print(f"\nSalvando resultados em: {arquivo_saida}")
        
        with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
            # Salvar o DataFrame no Excel
            df_pdvs.to_excel(writer, index=False, sheet_name="Sheet1")
            
            # Acessar a planilha para aplicar formatação
            workbook = writer.book
            worksheet = writer.sheets["Sheet1"]
            
            # Definir cores para a coluna PRIORIDADE
            fill_alta = PatternFill(start_color="ec6f50", end_color="ec6f50", fill_type="solid")  # Vermelho
            fill_media = PatternFill(start_color="ffe994", end_color="ffe994", fill_type="solid")  # Amarelo
            fill_baixa = PatternFill(start_color="77bc65", end_color="77bc65", fill_type="solid")  # Verde
            fill_offline = PatternFill(start_color="cccccc", end_color="cccccc", fill_type="solid")  # Cinza
            
            # Encontrar a coluna PRIORIDADE
            prioridade_col_idx = None
            for idx, col in enumerate(df_pdvs.columns, start=1):  # +1 porque openpyxl começa do 1, não do 0
                if col == "PRIORIDADE":
                    prioridade_col_idx = idx
                    break
            
            # Aplicar cores na coluna PRIORIDADE
            if prioridade_col_idx:
                for idx, row in enumerate(df_pdvs["PRIORIDADE"], start=2):  # Começa na linha 2 (cabeçalho está na linha 1)
                    cell = worksheet.cell(row=idx, column=prioridade_col_idx)
                    if row == "Alta":
                        cell.fill = fill_alta
                    elif row == "Média":
                        cell.fill = fill_media
                    elif row == "Offline":
                        cell.fill = fill_offline
                    elif row == "":
                        cell.fill = fill_baixa  # Células vazias (baixa prioridade) recebem cor verde
        
        # Estatísticas de sucesso
        sucessos = len([r for r in results.values() if r.get("STATUS") == "Sucesso"])
        print("\nProcesso concluído!")
        print(f"Total de PDVs processados: {len(df_pdvs)}")
        print(f"PDVs com sucesso: {sucessos}")
        print(f"PDVs com erro: {len(df_pdvs) - sucessos}")
        print(f"Verifique o arquivo: {arquivo_saida}")
    
    except Exception as e:
        logging.critical(f"Erro no processamento principal: {str(e)}")
        print(f"Ocorreu um erro: {str(e)}")
        print("Verifique o arquivo de log para mais detalhes.")
>>>>>>> bb3374c8ebbc2839fd2ae95848523d2e15acb503

if __name__ == "__main__":
    main()